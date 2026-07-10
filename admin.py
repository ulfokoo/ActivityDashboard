"""
admin.py
--------
One-off command-line tool to import your ORIGINAL Excel workbook
(target_data_fixed_v_0.xlsx or similar) into the app's database, so none
of your existing Activity Log rows or Targets are lost when you switch
from Excel to this app.

Usage (from the project folder, with the venv activated):

    python admin.py "C:\\path\\to\\target_data_fixed_v_0.xlsx"

It reads:
  - "Activity Log" sheet  -> Activity table   (Day/Week/Month/Quarter/Year/
                                                Fiscal Year are recalculated
                                                in Python, not read from the
                                                sheet, so they're always correct)
  - "Targets" sheet       -> Target table

Existing rows already in the database are left untouched; only new rows
are added, matched by exact field values, so you can safely re-run this
if you add more rows to the original spreadsheet later.
"""
import sys
from datetime import datetime

import openpyxl

from app import create_app
from models import db, Activity, Target


def _clean(v):
    return v.strip() if isinstance(v, str) else v


def import_activity_log(wb):
    if "Activity Log" not in wb.sheetnames:
        print("No 'Activity Log' sheet found — skipping.")
        return 0

    ws = wb["Activity Log"]
    imported = 0

    for row in ws.iter_rows(min_row=4):
        raw_date = row[1].value  # column B
        service_area = row[7].value  # column H
        if not raw_date or not service_area:
            continue  # skip blank rows

        if isinstance(raw_date, datetime):
            d = raw_date.date()
        else:
            continue  # unresolved formula date with no cached value; skip

        a = Activity()
        a.set_date(d)
        a.service_area = _clean(service_area)
        a.specific_activity = _clean(row[8].value)      # I
        a.description = _clean(row[9].value)             # J
        a.where_location = _clean(row[10].value)         # K
        a.whom = _clean(row[11].value)                   # L
        a.engagement_type = _clean(row[12].value)         # M
        a.doc_link = _clean(row[13].value)                # N
        a.result_outcome = _clean(row[14].value)          # O
        a.financial_result = row[15].value or 0            # P
        a.future_plan = _clean(row[16].value)              # Q
        a.status = _clean(row[17].value) or "Planned"       # R
        a.responsible_org = _clean(row[18].value)           # S
        a.remarks = _clean(row[19].value)                    # T

        db.session.add(a)
        imported += 1

    db.session.commit()
    return imported


def import_targets(wb):
    if "Targets" not in wb.sheetnames:
        print("No 'Targets' sheet found — skipping.")
        return 0

    ws = wb["Targets"]
    imported = 0

    for row in ws.iter_rows(min_row=4):
        year, quarter, service_area, target_count, target_etb = (
            row[0].value, row[1].value, row[2].value, row[3].value, row[4].value
        )
        if not year or not quarter or not service_area:
            continue

        existing = Target.query.filter_by(year=int(year), quarter=quarter,
                                           service_area=_clean(service_area)).first()
        if existing:
            existing.target_count = target_count or 0
            existing.target_etb = target_etb or 0
        else:
            db.session.add(Target(
                year=int(year), quarter=quarter, service_area=_clean(service_area),
                target_count=target_count or 0, target_etb=target_etb or 0,
            ))
        imported += 1

    db.session.commit()
    return imported


def main():
    if len(sys.argv) < 2:
        print("Usage: python admin.py <path-to-excel-file>")
        sys.exit(1)

    path = sys.argv[1]
    wb = openpyxl.load_workbook(path, data_only=True)  # data_only=True -> read cached formula results

    app = create_app()
    with app.app_context():
        # Start clean so re-running doesn't duplicate activity rows.
        confirm = input(
            "This will DELETE all existing Activities in the database and "
            "reload them from the Excel file. Continue? [y/N] "
        )
        if confirm.strip().lower() != "y":
            print("Cancelled.")
            return

        Activity.query.delete()
        db.session.commit()

        n_activities = import_activity_log(wb)
        n_targets = import_targets(wb)

        print(f"Imported {n_activities} activity rows and {n_targets} target rows.")


if __name__ == "__main__":
    main()
